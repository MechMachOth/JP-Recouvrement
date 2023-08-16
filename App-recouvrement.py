from collections import defaultdict
import tkinter as tk
from tkinter import filedialog
import openpyxl
import xlsxwriter
from datetime import datetime
from tkinter import messagebox
from tkinter import simpledialog, filedialog, ttk
from tkinter import *
from decimal import Decimal

g_rais = 0
g_date1 = 0
g_date2 = 0
g_nbr_fact = 0
g_ttc = 0
g_cré = 0
g_nbr_comm = 0
g_nbr_imp = 0
g_t_a_e = 0
g_client = 0
g_nbr_c_imp = 0
g_t_c_imp = 0
retour_recherch = 0
nom_rechercher = ''
ListCheq = []
ListFact = []
merged_list = []
changerowcolo = 0
data_fact = []
data_cheq = []


def read_data_from_excel(input_file_path):
    wb = openpyxl.load_workbook(input_file_path)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        # Crée un tableau pour chaque ligne et ajoute la ligne dans le tableau
        row_data = [cell for cell in row]
        data.append(row_data)

    return data


def ShowResult():
    global merged_list
    global changerowcolo
    global data_cheq
    global data_fact
    data_glob_téléch = []

    def téléch_glob():
        new_file_path = "Factures-impayées.xlsx"
        workbook = xlsxwriter.Workbook(new_file_path)
        fiche_impayé = workbook.add_worksheet('Factures impayées')
        fiche_impayé.set_row(0, 35)
        fiche_impayé.autofilter('A1:M11')
        fiche_impayé.set_zoom(86.5)

        format1 = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
        format1.set_bg_color('#00B0F0')
        format1.set_border()
        format1.set_border_color('#000000')
        format1.set_bold()
        format1.set_center_across()
        format1.set_shrink()
        format1.set_font_color('#44546A')
        format1.set_font_size(13)

        format2 = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
        format2.set_bg_color('#FFFFFF')
        format2.set_border()
        format2.set_border_color('#000000')
        format2.set_center_across()
        format2.set_shrink()
        format2.set_font_size(11)

        date_format = workbook.add_format(
            {'num_format': 'dd/mm/yyyy', 'align': 'center', 'valign': 'vcenter'})
        date_format.set_bg_color('#FFFFFF')
        date_format.set_border()
        date_format.set_border_color('#000000')
        date_format.set_center_across()
        date_format.set_shrink()
        date_format.set_font_size(11)

        fiche_impayé.write(0, 0, 'Raison sociale', format1)
        fiche_impayé.write(0, 2, 'Date première facture', format1)
        fiche_impayé.write(0, 3, 'Date dernière facture', format1)
        fiche_impayé.write(0, 1, "Ville", format1)
        fiche_impayé.write(0, 4, 'Nombre de facture', format1)
        fiche_impayé.write(0, 5, 'Total TTC', format1)
        fiche_impayé.write(0, 6, 'Créance', format1)
        fiche_impayé.write(0, 7, 'Nombre de facture commencée', format1)
        fiche_impayé.write(0, 8, 'Nombre de facture impayée', format1)
        fiche_impayé.write(0, 9, 'Total à encaisser', format1)
        fiche_impayé.write(0, 10, 'Client', format1)
        fiche_impayé.write(0, 11, 'Nombre chèque impayé', format1)
        fiche_impayé.write(0, 12, 'total chèque impayé', format1)

        fiche_impayé.set_column(2, 2, 20)
        fiche_impayé.set_column(0, 0, 46)
        fiche_impayé.set_column(1, 1, 20)
        fiche_impayé.set_column(3, 3, 20)
        fiche_impayé.set_column(4, 4, 20)
        fiche_impayé.set_column(5, 5, 20)
        fiche_impayé.set_column(6, 6, 20)
        fiche_impayé.set_column(7, 7, 20)
        fiche_impayé.set_column(8, 8, 20)
        fiche_impayé.set_column(9, 9, 20)
        fiche_impayé.set_column(10, 10, 20)
        fiche_impayé.set_column(11, 11, 20)
        fiche_impayé.set_column(12, 12, 20)

        for i in range(len(data_glob_téléch)):
            fiche_impayé.set_row(i+1, 25)
            fiche_impayé.write(i+1, 0, data_glob_téléch[i][0], format2)
            fiche_impayé.write(i+1, 1, data_glob_téléch[i][1], format2)
            fiche_impayé.write(i+1, 2, data_glob_téléch[i][2], date_format)
            fiche_impayé.write(i+1, 3, data_glob_téléch[i][3], date_format)
            fiche_impayé.write(i+1, 4, data_glob_téléch[i][4], format2)
            fiche_impayé.write(i+1, 5, data_glob_téléch[i][5], format2)
            fiche_impayé.write(i+1, 6, data_glob_téléch[i][6], format2)
            fiche_impayé.write(i+1, 7, data_glob_téléch[i][7], format2)
            fiche_impayé.write(i+1, 8, data_glob_téléch[i][8], format2)
            fiche_impayé.write(i+1, 9, data_glob_téléch[i][9], format2)
            fiche_impayé.write(i+1, 10, data_glob_téléch[i][10], format2)
            fiche_impayé.write(i+1, 11, data_glob_téléch[i][11], format2)
            fiche_impayé.write(i+1, 12, data_glob_téléch[i][12], format2)

        workbook.close()
        messagebox.showinfo("Traitement terminé",
                            f'Nouveau fichier "{new_file_path}" a été créé')

    def rowscolor():
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        global changerowcolo
        global merged_list
        mytag = ''
        for parent in my_tree.get_children():
            my_tree.delete(parent)
        if changerowcolo == 0:
            changerowcolo = 1
            for i in merged_list:
                if mytag == 'normal':
                    mytag = 'gray'
                else:
                    mytag = 'normal'
                my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]), tags=(mytag))
        else:
            changerowcolo = 0
            mytag = 'normal'
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text=i, values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]), tags=(mytag))
    if ListFact:
        cheq_dict = {item[0]: item for item in ListCheq}

        for fact in ListFact:
            entreprise = fact[0]
            if entreprise in cheq_dict:
                cheq_info = cheq_dict[entreprise]
                merged_item = fact + cheq_info[1:]
            else:
                merged_item = fact + ['--', '--', '--']
            merged_list.append(merged_item)

    for cheq in ListCheq:
        entreprise = cheq[0]
        if entreprise not in [item[0] for item in merged_list]:
            merged_item = [entreprise] + ['--', '--', '--',
                                          '--', '--', '--', '--', '--', '--'] + cheq[1:]
            merged_list.append(merged_item)
    label.pack_forget()
    btn_valider.pack_forget()
    checkbox_frame.pack_forget()
    root.resizable(width=1, height=1)
    screen_width = int(root.winfo_screenwidth())
    screen_height = int(root.winfo_screenheight())
    root.geometry(f"{screen_width}x{screen_height}")

    # Premier cadre (au-dessus du deuxième cadre)
    frame_top = Frame(root, background='black')
    frame_top.grid(row=0, column=0, sticky="nsew")

    # Deuxième cadre (en bas)
    frame_bottom = Frame(root, background='black')
    frame_bottom.grid(row=1, column=0, sticky="nsew")

    # Configurer le redimensionnement des lignes et colonnes
    root.grid_rowconfigure(0, weight=9)  # Prend 9/10 de l'espace vertical
    root.grid_rowconfigure(1, weight=1)  # Prend 1/10 de l'espace vertical
    root.grid_columnconfigure(0, weight=1)  # Prend tout l'espace horizontal


    style = ttk.Style()
    style.theme_use('default')
    style.configure("Treeview", foreground="black",
                    fieldbackground="silver", rowheight=25)
    scrolly = ttk.Scrollbar(frame_top, orient=VERTICAL)
    my_tree = ttk.Treeview(
        frame_top, height=37, yscrollcommand=scrolly.set)

    my_tree.tag_configure('gray', background='gray')
    my_tree.tag_configure('normal', background='white')
    my_tree.tag_configure('blue', background='lightblue')
    my_tree.tag_configure('green', background='lightgreen')
    my_tree.tag_configure('red', background='red')

    my_tree['columns'] = ("Raison sociale", "Date première facture", "Date dernière facture",
                          "Nbr Facture", "Total TTC", "Créance", "Nbr facture commencée", "Nbr facture impayée", "Total à encaisser", "Client", "Nbr chèque impayé", "Total chèque impayé")

    my_tree.column("#0", width=0, stretch=NO)
    my_tree.column("Raison sociale", width=240, anchor=CENTER, minwidth=25)
    my_tree.column("Date première facture", width=180,
                   anchor=CENTER, minwidth=25)
    my_tree.column("Date dernière facture", width=180,
                   anchor=CENTER, minwidth=25)
    my_tree.column("Nbr Facture", width=100,
                   anchor=CENTER, minwidth=25)
    my_tree.column("Total TTC", width=100, anchor=CENTER, minwidth=25)
    my_tree.column("Créance", width=100,
                   anchor=CENTER, minwidth=25)
    my_tree.column("Nbr facture commencée", width=190,
                   anchor=CENTER, minwidth=25)
    my_tree.column("Nbr facture impayée", width=170,
                   anchor=CENTER, minwidth=25)
    my_tree.column("Total à encaisser", width=150,
                   anchor=CENTER, minwidth=25)
    my_tree.column("Client", width=150,
                   anchor=CENTER, minwidth=25)
    my_tree.column("Nbr chèque impayé", width=150,
                   anchor=CENTER, minwidth=25)
    my_tree.column("Total chèque impayé", width=160,
                   anchor=CENTER, minwidth=25)

    my_tree.heading("#0", text="", anchor=W)
    my_tree.heading("Raison sociale", text="Raison sociale", anchor=CENTER)
    my_tree.heading("Date première facture",
                    text="Date première facture", anchor=CENTER)
    my_tree.heading("Date dernière facture",
                    text="Date dernière facture", anchor=CENTER)
    my_tree.heading("Nbr Facture", text="Nbr Facture", anchor=CENTER)
    my_tree.heading("Total TTC", text="Total TTC", anchor=CENTER)
    my_tree.heading("Créance", text="Créance", anchor=CENTER)
    my_tree.heading("Nbr facture commencée",
                    text="Nbr facture commencée", anchor=CENTER)
    my_tree.heading("Nbr facture impayée",
                    text="Nbr facture impayée", anchor=CENTER)
    my_tree.heading("Total à encaisser",
                    text="Total à encaisser", anchor=CENTER)
    my_tree.heading("Client",
                    text="Client", anchor=CENTER)
    my_tree.heading("Nbr chèque impayé",
                    text="Nbr chèque impayé", anchor=CENTER)
    my_tree.heading("Total chèque impayé",
                    text="Total chèque impayé", anchor=CENTER)

    my_tree.pack(pady=30, padx=20)
    tot_entre = 0
    tot_nbr_fac = 0
    tot_ttc = 0
    tot_cré = 0
    tot_nbr_fac_comm = 0
    tot_nbr_fac_imp = 0
    tot_a_enc = 0
    tot_tot_cheq_imp = 0
    tot_nbr_cheq_imp = 0
    for i in merged_list:
        tot_entre = tot_entre+1
        if i[2] != '--':
            tot_nbr_fac = tot_nbr_fac + int(i[2])

        if i[3] != '--':
            tot_ttc = tot_ttc + int(i[3])

        if i[6] != '--':
            tot_cré = tot_cré + int(i[6])

        if i[7] != '--':
            tot_nbr_fac_comm = tot_nbr_fac_comm + int(i[7])

        if i[8] != '--':
            tot_nbr_fac_imp = tot_nbr_fac_imp + int(i[8])

        if i[9] != '--':
            tot_a_enc = tot_a_enc + int(i[9])

        if i[11] != '--':
            tot_nbr_cheq_imp = tot_nbr_cheq_imp + int(i[10])

        if i[12] != '--':
            tot_tot_cheq_imp = tot_tot_cheq_imp + int(i[12])
        my_tree.insert(parent='', index='end', iid=i, text='', values=(
            i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        data_glob_téléch.append(
            [i[0], i[1], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]])

    scrolly.configure(command=my_tree.yview)
    scrolly.place(y=70, height=860, x=3)

    def orga_rais():
        global g_rais
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[0] == '--':
                return 'Z'
            return item[0]
        if g_rais == 0:
            g_rais = 1
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_rais = 0
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_date1():
        global g_date1
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[4] == '--':
                return -999999999
            return item[4]
        if g_date1 == 0:
            g_date1 = 1
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_date1 = 0
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()

        rais.configure(bg='black', fg='white')
        date1.configure(bg='white', fg='black')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_date2():
        global g_date2
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[5] == '--':
                return -999999999
            return item[5]
        if g_date2 == 0:
            g_date2 = 1
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_date2 = 0
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        date2.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_nbr_fact():
        global g_nbr_fact
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[2] == '--':
                return -999999999
            return item[2]
        if g_nbr_fact == 0:
            g_nbr_fact = 1
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_nbr_fact = 0
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        nbr_fact.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_ttc():
        global g_ttc
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[3] == '--':
                return -999999999
            return item[3]
        if g_ttc == 0:
            g_ttc = 1
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_ttc = 0
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        ttc.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_cré():
        global g_cré
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[6] == '--':
                return -999999999
            return item[6]
        if g_cré == 0:
            g_cré = 1
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_cré = 0
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        cré.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_nbr_comm():
        global g_nbr_comm
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[7] == '--':
                return -999999999
            return item[7]
        if g_nbr_comm == 0:
            g_nbr_comm = 1
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_nbr_comm = 0
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        nbr_comm.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_nbr_imp():
        global g_nbr_imp
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[8] == '--':
                return -999999999
            return item[8]
        if g_nbr_imp == 0:
            g_nbr_imp = 1
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_nbr_imp = 0
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        nbr_imp.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_t_a_e():
        global merged_list
        global g_t_a_e
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[9] == '--':
                return -999999999
            return item[9]
        if g_t_a_e == 0:
            g_t_a_e = 1
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_t_a_e = 0
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        t_a_e.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_client():
        global g_client
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[11] == '--':
                return 'Z'
            return item[11]
        if g_client == 0:
            g_client = 1

            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_client = 0
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        client.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_nbr_c_imp():
        global g_nbr_c_imp
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[10] == '--':
                return -999999999
            return item[10]
        if g_nbr_c_imp == 0:
            g_nbr_c_imp = 1
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_nbr_c_imp = 0
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        t_c_imp.configure(bg='black', fg='white')

    def orga_t_c_imp():
        global g_t_c_imp
        global merged_list
        global changerowcolo
        global retour_recherch
        global re_recherch
        if retour_recherch == 1:
            retour_recherch = 0
            re_recherche.place_forget()
        for parent in my_tree.get_children():
            my_tree.delete(parent)

        def custom_sort_key(item):
            if item[12] == '--':
                return -999999999
            return item[12]
        if g_t_c_imp == 0:
            g_t_c_imp = 1
            merged_list_sorted = sorted(
                merged_list, key=custom_sort_key, reverse=True)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        else:
            g_t_c_imp = 0
            merged_list_sorted = sorted(merged_list, key=custom_sort_key)
            merged_list = merged_list_sorted
            for i in merged_list:
                my_tree.insert(parent='', index='end', iid=i, text='', values=(
                    i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
        if changerowcolo == 0:
            changerowcolo = 1
            rowscolor()
        else:
            changerowcolo = 0
            rowscolor()
        rais.configure(bg='black', fg='white')
        t_c_imp.configure(bg='white', fg='black')
        date1.configure(bg='black', fg='white')
        date2.configure(bg='black', fg='white')
        nbr_fact.configure(bg='black', fg='white')
        ttc.configure(bg='black', fg='white')
        cré.configure(bg='black', fg='white')
        nbr_comm.configure(bg='black', fg='white')
        nbr_imp.configure(bg='black', fg='white')
        t_a_e.configure(bg='black', fg='white')
        client.configure(bg='black', fg='white')
        nbr_c_imp.configure(bg='black', fg='white')

    def show_detail(tree):
        global data_cheq
        global data_fact
        détail_fact = []
        détail_chèque = []
        data_fact_téléch = []
        data_chèque_téléch = []

        def téléch_détail():
            new_file_path = "Factures-impayées-"+values[0]+".xlsx"
            workbook = xlsxwriter.Workbook(new_file_path)
            fiche_impayé_détail = workbook.add_worksheet('Factures impayées')
            fiche_impayé_détail.set_row(0, 35)
            fiche_impayé_détail.set_row(1, 35)
            fiche_impayé_détail.filter_column_list('A2', 'F2')
            fiche_impayé_détail.filter_column_list('H2', 'M2')
            fiche_impayé_détail.set_zoom(99)
            format_centre_gras_souligné = workbook.add_format({
                'bold': True,
                'underline': True,
                'align': 'center',
                'valign': 'vcenter'
            })

            format_centre_gras_souligné.set_font_size(20)
            fiche_impayé_détail.merge_range(
                'A1:F1', 'Facture(s)', format_centre_gras_souligné)
            fiche_impayé_détail.merge_range(
                'H1:L1', 'Chèque(s)', format_centre_gras_souligné)

            format1 = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter'})
            format1.set_bg_color('#00B0F0')
            format1.set_border()
            format1.set_border_color('#000000')
            format1.set_bold()
            format1.set_center_across()
            format1.set_shrink()
            format1.set_font_color('#44546A')
            format1.set_font_size(13)

            format2 = workbook.add_format(
                {'align': 'center', 'valign': 'vcenter'})
            format2.set_bg_color('#FFFFFF')
            format2.set_border()
            format2.set_border_color('#000000')
            format2.set_center_across()
            format2.set_shrink()
            format2.set_font_size(11)

            date_format = workbook.add_format(
                {'num_format': 'dd/mm/yyyy', 'align': 'center', 'valign': 'vcenter'})
            date_format.set_bg_color('#FFFFFF')
            date_format.set_border()
            date_format.set_border_color('#000000')
            date_format.set_center_across()
            date_format.set_shrink()
            date_format.set_font_size(11)

            fiche_impayé_détail.write(1, 0, 'Réf facture', format1)
            fiche_impayé_détail.write(1, 2, 'Total TTC', format1)
            fiche_impayé_détail.write(1, 3, 'Créance', format1)
            fiche_impayé_détail.write(1, 1, "Date facture", format1)
            fiche_impayé_détail.write(1, 4, 'Facture commencée', format1)
            fiche_impayé_détail.write(1, 5, 'total à encaisser', format1)

            fiche_impayé_détail.write(1, 7, 'Réf facture', format1)
            fiche_impayé_détail.write(1, 8, 'Date paiement', format1)
            fiche_impayé_détail.write(1, 9, 'Type paiement', format1)
            fiche_impayé_détail.write(1, 10, 'Montant paiement', format1)
            fiche_impayé_détail.write(1, 11, 'Etat paiement', format1)

            fiche_impayé_détail.set_column(2, 2, 20)
            fiche_impayé_détail.set_column(0, 0, 20)
            fiche_impayé_détail.set_column(1, 1, 20)
            fiche_impayé_détail.set_column(3, 3, 20)
            fiche_impayé_détail.set_column(4, 4, 20)
            fiche_impayé_détail.set_column(5, 5, 20)

            fiche_impayé_détail.set_column(7, 7, 20)
            fiche_impayé_détail.set_column(8, 8, 20)
            fiche_impayé_détail.set_column(9, 9, 20)
            fiche_impayé_détail.set_column(10, 10, 20)
            fiche_impayé_détail.set_column(11, 11, 40)

            for i in range(len(data_fact_téléch)):
                row_index = i + 2
                fiche_impayé_détail.set_row(row_index, 25)
                fiche_impayé_détail.write(
                    row_index, 0, data_fact_téléch[i][0], format2)
                fiche_impayé_détail.write(
                    row_index, 1, data_fact_téléch[i][1], date_format)
                fiche_impayé_détail.write(
                    row_index, 2, data_fact_téléch[i][2], format2)
                fiche_impayé_détail.write(
                    row_index, 3, data_fact_téléch[i][3], format2)
                fiche_impayé_détail.write(
                    row_index, 4, data_fact_téléch[i][4], format2)
                fiche_impayé_détail.write(
                    row_index, 5, data_fact_téléch[i][5], format2)

            for i in range(len(data_chèque_téléch)):
                row_index = i + 2
                fiche_impayé_détail.set_row(row_index, 25)
                fiche_impayé_détail.write(
                    row_index, 7, data_chèque_téléch[i][0], format2)
                fiche_impayé_détail.write(
                    row_index, 8, data_chèque_téléch[i][1], date_format)
                fiche_impayé_détail.write(
                    row_index, 9, data_chèque_téléch[i][2], format2)
                fiche_impayé_détail.write(
                    row_index, 10, data_chèque_téléch[i][3], format2)
                fiche_impayé_détail.write(
                    row_index, 11, data_chèque_téléch[i][4], format2)
            workbook.close()
            messagebox.showinfo("Traitement terminé",
                                f'Nouveau fichier "{new_file_path}" a été créé')
        curItem = tree.focus()
        values = tree.item(curItem, "values")

        for entry in data_fact:
            if entry[4].upper() == values[0]:
                difference = entry[7] - entry[11]
                if entry[7] == entry[11]:
                    entry.append('NON')
                elif entry[7] > entry[11] :
                    entry.append('OUI')
                formatted_difference = "{:.2f}".format(difference)
                entry.append(formatted_difference)
                détail_fact.append(entry)

        for entry in data_cheq:
            if entry[3].upper() == values[0]:
                détail_chèque.append(entry)

        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()

        topDetail = Toplevel(root)
        topDetail.title(values[0])

        top_width = int(screen_width)
        top_height = int(screen_height * 0.5)

        x_coordinate = 0
        y_coordinate = int((screen_height - top_height) / 2)

        topDetail.geometry(
            f"{top_width}x{top_height}+{x_coordinate}+{y_coordinate}")


        frame_fact = Frame(topDetail)
        frame_fact.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        frame_cheque = Frame(topDetail)
        frame_cheque.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        frame_bottom = Frame(topDetail)
        frame_bottom.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

        # Configurer le redimensionnement des lignes et colonnes pour les cadres et les widgets internes
        topDetail.grid_rowconfigure(0, weight=1)
        topDetail.grid_columnconfigure(0, weight=1)
        topDetail.grid_columnconfigure(1, weight=1)

        frame_fact.grid_rowconfigure(0, weight=1)
        frame_fact.grid_columnconfigure(0, weight=1)

        frame_cheque.grid_rowconfigure(0, weight=1)
        frame_cheque.grid_columnconfigure(0, weight=1)

        frame_bottom.grid_rowconfigure(0, weight=1)
        frame_bottom.grid_columnconfigure(0, weight=1)


        label_fact = Label(frame_fact, text="Facture(s)", font=(
            'Times', 15), fg='black', highlightbackground="black", highlightthickness=2)
        label_fact.pack(side=TOP)

        label_cheque = Label(frame_cheque, text="Chèque(s)", font=(
            'Times', 15), fg='black', highlightbackground="black", highlightthickness=2)
        label_cheque.pack(side=TOP)

        style = ttk.Style()
        style.theme_use('default')
        style.configure("Treeview", foreground="black",
                        fieldbackground="silver", rowheight=25)
        scrolly_detail = ttk.Scrollbar(frame_fact, orient=VERTICAL)
        my_tree_detail_fact = ttk.Treeview(
            frame_fact, height=37, yscrollcommand=scrolly_detail.set)

        my_tree_detail_fact['columns'] = (
            'Réf facture', "Date facture", "Total TTC", "Créance", "Facture commencée",  "Total à encaisser")

        my_tree_detail_fact.column("#0", width=0, stretch=NO)
        my_tree_detail_fact.column(
            "Réf facture", width=200, anchor=CENTER, minwidth=25)
        my_tree_detail_fact.column("Date facture", width=180,
                                   anchor=CENTER, minwidth=25)
        my_tree_detail_fact.column("Total TTC", width=120,
                                   anchor=CENTER, minwidth=25)
        my_tree_detail_fact.column("Créance", width=100,
                                   anchor=CENTER, minwidth=25)
        my_tree_detail_fact.column("Facture commencée", width=190,
                                   anchor=CENTER, minwidth=25)
        my_tree_detail_fact.column("Total à encaisser", width=150,
                                   anchor=CENTER, minwidth=25)

        my_tree_detail_fact.heading("#0", text="", anchor=W)
        my_tree_detail_fact.heading(
            "Réf facture", text="Réf facture", anchor=CENTER)
        my_tree_detail_fact.heading("Date facture",
                                    text="Date facture", anchor=CENTER)
        my_tree_detail_fact.heading("Total TTC",
                                    text="Total TTC", anchor=CENTER)
        my_tree_detail_fact.heading("Créance", text="Créance", anchor=CENTER)
        my_tree_detail_fact.heading("Facture commencée",
                                    text="Facture commencée", anchor=CENTER)
        my_tree_detail_fact.heading("Total à encaisser",
                                    text="Total à encaisser", anchor=CENTER)

        for i in détail_fact:
            print(i)
            my_tree_detail_fact.insert(parent='', index='end', iid=i, text='', values=(
                i[2], i[0], i[7], i[11], i[27], i[28]))
            data_fact_téléch.append([i[2], i[0], i[7], i[11], i[27], i[28]])

        my_tree_detail_fact.pack(pady=30, padx=20)
        scrolly_detail.configure(command=my_tree_detail_fact.yview)
        scrolly_detail.place(y=20, height=500, x=3)

        scrolly_detail_chèque = ttk.Scrollbar(frame_cheque, orient=VERTICAL)
        my_tree_detail_chèque = ttk.Treeview(
            frame_cheque, height=37, yscrollcommand=scrolly_detail_chèque.set)

        my_tree_detail_chèque['columns'] = (
            'Réf facture', "Date paiement", "Type paiement", "Montant paiement", "Etat paiement")

        my_tree_detail_chèque.column("#0", width=0, stretch=NO)
        my_tree_detail_chèque.column(
            "Réf facture", width=200, anchor=CENTER, minwidth=25)
        my_tree_detail_chèque.column("Date paiement", width=180,
                                     anchor=CENTER, minwidth=25)
        my_tree_detail_chèque.column("Type paiement", width=150,
                                     anchor=CENTER, minwidth=25)
        my_tree_detail_chèque.column("Montant paiement", width=100,
                                     anchor=CENTER, minwidth=25)
        my_tree_detail_chèque.column("Etat paiement", width=190,
                                     anchor=CENTER, minwidth=25)

        my_tree_detail_chèque.heading("#0", text="", anchor=W)
        my_tree_detail_chèque.heading(
            "Réf facture", text="Réf facture", anchor=CENTER)
        my_tree_detail_chèque.heading("Date paiement",
                                      text="Date paiement", anchor=CENTER)
        my_tree_detail_chèque.heading("Type paiement",
                                      text="Type paiement", anchor=CENTER)
        my_tree_detail_chèque.heading(
            "Montant paiement", text="Montant paiement", anchor=CENTER)
        my_tree_detail_chèque.heading("Etat paiement",
                                      text="Etat paiement", anchor=CENTER)

        for i in détail_chèque:
            if i[11] == "Chèque":
                my_tree_detail_chèque.insert(parent='', index='end', iid=i, text='', values=(
                    i[1], i[8], i[11], i[6], i[12]))
                data_chèque_téléch.append([i[1], i[8], i[11], i[6], i[12]])

        my_tree_detail_chèque.pack(pady=30, padx=20)
        scrolly_detail_chèque.configure(command=my_tree_detail_chèque.yview)
        scrolly_detail_chèque.place(y=20, height=500, x=3)

        téléch_xlsx_détail = tk.Button(
            frame_bottom, text="Télécharger⬇️", command=téléch_détail)
        téléch_xlsx_détail.pack()
        topDetail.bind("<Escape>", lambda e: topDetail.destroy())
        topDetail.bind("<Escape>", lambda e: topDetail.destroy())

    def searchNam():
        topSearch = Toplevel()
        topSearch.title("Recherche")
        topSearch.geometry("1000x370")

        def stopSearch():
            topSearch.destroy()

        def tl31(e):
            global merged_list
            typed = e1.get()
            if typed == '':
                for parent in merged_list:
                    loubil.insert(END, parent[0])
            else:
                loubil.delete(0, END)
                for parent in merged_list:
                    if typed.upper() in parent[0].upper():
                        loubil.insert(END, parent[0])
            e1.focus_set()

        def autofill(event):
            e1.delete(0, END)
            selected_item = loubil.get(loubil.curselection())
            e1.insert(0, selected_item)
            tl31(None)

        def searchData():
            global retour_recherch
            global merged_list
            global nom_rechercher
            global changerowcolo
            global re_recherche

            def reload_data():
                global changerowcolo
                global retour_recherch
                global re_recherche
                retour_recherch = 0
                re_recherche.place_forget()
                for parent in my_tree.get_children():
                    my_tree.delete(parent)
                for i in merged_list:
                    my_tree.insert(parent='', index='end', iid=i, text='', values=(
                        i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))
                if changerowcolo == 0:
                    changerowcolo = 1
                    rowscolor()
                else:
                    changerowcolo = 0
                    rowscolor()

            if retour_recherch == 0:
                re_recherche = tk.Button(
                    text='<-', bg='black', fg='white', command=reload_data)
                re_recherche.place(x=10, y=980)
                retour_recherch = 1
                root.bind("<Control-Z>", lambda e: reload_data())
                root.bind("<Control-z>", lambda e: reload_data())

            nom_rechercher = e1.get().upper()
            e1.delete(0, END)
            stopSearch()
            retour_recherch = 1
            for parent in my_tree.get_children():
                my_tree.delete(parent)
            for i in merged_list:
                if nom_rechercher in i[0]:
                    my_tree.insert(parent='', index='end', iid=i, text='', values=(
                        i[0], i[4], i[5], i[2], i[3], i[6], i[7], i[8], i[9], i[11], i[10], i[12]))

        loubil = Listbox(topSearch, bg='white',
                         activestyle='dotbox', justify="center")
        loubil.place(width=550, height=300, x=430, y=30)

        loubil.bind("<<ListboxSelect>>", autofill)
        topSearch.resizable(width=0, height=0)
        loubil.delete(0, END)
        c1 = tk.Label(topSearch, text='   Raison Sociale                                          ',
                      font=('Times', 11, 'bold'))
        c1.place(x=70, y=155)
        e1 = Entry(
            topSearch, textvariable="", width=25)
        e1.place(x=200, y=150)
        e1.focus_set()
        e1.bind("<KeyRelease>", tl31)
        ch = Button(
            topSearch, text='Chercher', )
        ch.configure(
            font=('Times', 11, 'bold'), bg='green', fg='white', command=searchData)
        ch.place(x=300, y=200)
        anul = Button(
            topSearch, text='Annulé')
        anul.configure(
            font=('Times', 11, 'bold'), bg='red', fg='white', command=stopSearch)
        anul.place(x=200, y=200)
        topSearch.protocol("WM_DELETE_WINDOW", stopSearch)
        topSearch.bind(
            "<Return>", lambda e: searchData())
        topSearch.bind(
            "<Escape>", lambda e: stopSearch())

    rais = tk.Button(frame_top, text=f"{tot_entre} ⤋", highlightbackground="black", fg='white',
                     background='black', command=orga_rais)
    rais.place(y=0, x=120)

    date1 = tk.Button(frame_top, text="⤋", highlightbackground="black", fg='white',
                      background='black', command=orga_date1)
    date1.place(y=0, x=350)

    date2 = tk.Button(frame_top, text="⤋", highlightbackground="black", fg='white',
                      background='black', command=orga_date2)
    date2.place(y=0, x=530)

    nbr_fact = tk.Button(frame_top, text=f"{tot_nbr_fac} ⤋", highlightbackground="black", fg='white',
                         background='black', command=orga_nbr_fact)
    nbr_fact.place(y=0, x=645)

    ttc = tk.Button(frame_top, text=f"{tot_ttc} ⤋", highlightbackground="black", fg='white',
                    background='black', command=orga_ttc)
    ttc.place(y=0, x=720)

    cré = tk.Button(frame_top, text=f"{tot_cré} ⤋", highlightbackground="black", fg='black',
                    background='white', command=orga_cré)
    cré.place(y=0, x=820)

    nbr_comm = tk.Button(frame_top, text=f"{tot_nbr_fac_comm} ⤋", highlightbackground="black", fg='white',
                         background='black', command=orga_nbr_comm)
    nbr_comm.place(y=0, x=995)

    nbr_imp = tk.Button(frame_top, text=f"{tot_nbr_fac_imp} ⤋", highlightbackground="black", fg='white',
                        background='black', command=orga_nbr_imp)
    nbr_imp.place(y=0, x=1185)

    t_a_e = tk.Button(frame_top, text=f"{tot_a_enc} ⤋", highlightbackground="black", fg='white',
                      background='black', command=orga_t_a_e)
    t_a_e.place(y=0, x=1320)

    client = tk.Button(frame_top, text="⤋", highlightbackground="black", fg='white',
                       background='black', command=orga_client)
    client.place(y=0, x=1495)

    nbr_c_imp = tk.Button(frame_top, text=f"{tot_nbr_cheq_imp} ⤋", highlightbackground="black", fg='white',
                          background='black', command=orga_nbr_c_imp)
    nbr_c_imp.place(y=0, x=1630)

    t_c_imp = tk.Button(frame_top, text=f"{tot_tot_cheq_imp} ⤋", highlightbackground="black", fg='white',
                        background='black', command=orga_t_c_imp)
    t_c_imp.place(y=0, x=1775)

    rowscolo = tk.Button(frame_top, text="•", highlightbackground="black", fg='white',
                         background='gray', command=rowscolor)
    rowscolo.place(y=0, x=0)


    frame_left = Frame(frame_bottom, background='gray')
    frame_right = Frame(frame_bottom, background='gray')

    frame_bottom.grid_rowconfigure(0, weight=1)  # Assurez-vous que la ligne occupe tout l'espace vertical disponible
    frame_bottom.grid_columnconfigure(0, weight=1)  # Assurez-vous que la colonne occupe tout l'espace horizontal disponible

    frame_left.grid(row=0, column=0, padx=10, pady=10, sticky="w")
    frame_right.grid(row=0, column=1, padx=10, pady=10, sticky="e")

    search = tk.Button(frame_left, text="Recherche", command=searchNam)
    search.pack(side="left")

    téléch_xlsx_glob = tk.Button(frame_right, text="Télécharger", command=téléch_glob)
    téléch_xlsx_glob.pack(side="right")

    root.bind("<Control-f>", lambda e: searchNam())
    root.bind("<Control-F>", lambda e: searchNam())
    my_tree.bind("<Double-1>", lambda e: show_detail(my_tree))
    my_tree.bind("<Return>", lambda e: show_detail(my_tree))


def TraitListFact():
    global ListFact
    global data_fact
    file_path_fact = filedialog.askopenfilename(
        title="Select file 'Liste factures impayées'", filetypes=[("Fichiers Excel", "*.xlsx")])
    if file_path_fact:
        data = read_data_from_excel(file_path_fact)
        data_fact = data

        entreprise_info = defaultdict(int)
        entreprise_ville = {}
        entreprise_dates = defaultdict(list)
        entreprise_total = defaultdict(Decimal)
        entreprise_cumule_creance = defaultdict(Decimal)
        cumule_inf_count = defaultdict(int)
        cumule_eq_count = defaultdict(int)
        entete = 0
        for entry in data:
            if entete != 0:
                nom_entreprise = entry[4]
                ville_entreprise = entry[6]
                date_facture = entry[0]
                cumule_entreprise_str = str(entry[7])
                cumule_creance_str = str(entry[11])

                cumule_entreprise_str = cumule_entreprise_str.replace(",", ".")
                cumule_creance_str = cumule_creance_str.replace(",", ".")
                cumule_entreprise = Decimal(cumule_entreprise_str)
                cumule_creance = Decimal(cumule_creance_str)

                entreprise_info[nom_entreprise] += 1
                entreprise_ville[nom_entreprise] = ville_entreprise
                entreprise_total[nom_entreprise] += cumule_entreprise
                entreprise_dates[nom_entreprise].append(date_facture)
                entreprise_cumule_creance[nom_entreprise] += cumule_creance

                if cumule_creance < cumule_entreprise:
                    cumule_inf_count[nom_entreprise] += 1
                elif cumule_creance == cumule_entreprise:
                    cumule_eq_count[nom_entreprise] += 1

            else:
                entete = 1

        for nom, count in entreprise_info.items():
            cumule_creance = entreprise_cumule_creance[nom]
            cumule_entreprise = entreprise_total[nom]
            cumule_inf = cumule_inf_count[nom]
            cumule_eq = cumule_eq_count[nom]
            premiere_facture = min(entreprise_dates[nom])
            derniere_facture = max(entreprise_dates[nom])

            total_a_encaisser = cumule_entreprise - \
                cumule_creance if cumule_creance < cumule_entreprise else 0

            ligne_resultat = [nom.upper(), entreprise_ville[nom], count, cumule_entreprise,
                              premiere_facture, derniere_facture, cumule_creance, cumule_inf, cumule_eq,
                              total_a_encaisser]
            ListFact.append(ligne_resultat)
        ListFact_sorted = sorted(ListFact, key=lambda x: x[6], reverse=True)
        ListFact = ListFact_sorted


def TraitListCheq():
    global ListCheq
    global data_cheq
    file_path_cheq = filedialog.askopenfilename(
        title="Select file 'Liste règlements des factures'", filetypes=[("Fichiers Excel", "*.xlsx")])
    if file_path_cheq:
        data = read_data_from_excel(file_path_cheq)
        data_cheq = data

        # Dictionnaire pour stocker les informations sur chaque entreprise
        entreprise_info = defaultdict(lambda: {
                                      'cheques_impayes': 0, 'ch_client_ne_fait_plus_partie': 'OUI', 'total_montant_cheques': 0})

        for row in data:
            entreprise = row[3]
            montant = row[6]
            statut = row[12]

            if statut is not None:
                if "impayé" in statut.lower():
                    entreprise_info[entreprise]['cheques_impayes'] += 1
                    if isinstance(montant, (int, float)):
                        entreprise_info[entreprise]['total_montant_cheques'] += montant

                if "client ne fait plus partie" in statut.lower():
                    entreprise_info[entreprise]['ch_client_ne_fait_plus_partie'] = 'NON'

        for entreprise, info in entreprise_info.items():
            ListCheq.append([
                entreprise.upper(),
                info['cheques_impayes'],
                info['ch_client_ne_fait_plus_partie'],
                info['total_montant_cheques']
            ])


def afficher_selection():
    root.unbind("<Return>")
    selectionsFiches = []
    for i, var in enumerate(checkbox_vars):
        if var.get() == 1:
            selectionsFiches.append(fiches[i])

    if selectionsFiches == []:
        messagebox.showinfo("!!!!!", 'Aucun fichier selectionné')
        return

    for element in selectionsFiches:
        if element == "Liste factures impayées":
            TraitListFact()
        if element == "Liste règlements des factures":
            TraitListCheq()
    ShowResult()


def toggle_colors():
    for i, var in enumerate(checkbox_vars):
        if var.get() == 1:
            checkboxes[i].config(fg='black', bg='white')
        else:
            checkboxes[i].config(fg='white', bg='black')


if __name__ == "__main__":
    root = tk.Tk()
    root.title("  JP&CO")
    root.config(background='black')
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_cordinate = int((screen_width/2) - 160)
    y_cordinate = int((screen_height/2) - 140)
    root.geometry("{}x{}+{}+{}".format(400, 250, x_cordinate, y_cordinate))
    root.resizable(width=0, height=0)

    fiches = ["Liste factures impayées", "Liste règlements des factures"]

    label = tk.Label(root, text="Sélectionnez les fichiers à traiter",
                     font=('Bold', 15), fg='white', bg='black')
    label.pack(pady=20)

    checkbox_vars = []
    checkboxes = []

    checkbox_frame = tk.Frame(root, bg="black")
    checkbox_frame.pack()

    for i, fiche in enumerate(fiches):
        var = IntVar()
        checkbox = tk.Checkbutton(checkbox_frame, text=fiche, variable=var,
                                  font=('Bold', 15), fg='white', bg='black', command=toggle_colors)
        checkbox.pack(pady=10)
        checkbox_vars.append(var)
        checkboxes.append(checkbox)

    btn_valider = tk.Button(root, text="Valider", command=afficher_selection)
    btn_valider.pack(pady=20)
    root.bind("<Return>", lambda e: afficher_selection())

    root.mainloop()
